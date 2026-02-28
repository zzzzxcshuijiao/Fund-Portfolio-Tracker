"""Excel parser for 基金E账户App exported Excel files.

Expected format:
- Row 1: Title
- Row 2-3: User info
- Row 4: Empty
- Row 5: Column headers
- Row 6+: Data rows
"""

import hashlib
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Optional

import openpyxl

# Expected column mapping (1-indexed)
COLUMN_MAP = {
    "序号": 1,
    "基金代码": 2,
    "基金名称": 3,
    "份额类别": 4,
    "基金管理人": 5,
    "基金账户": 6,
    "销售机构": 7,
    "交易账户": 8,
    "持有份额": 9,
    "份额日期": 10,
    "基金净值": 11,
    "净值日期": 12,
    "资产情况": 13,
    "结算币种": 14,
    "分红方式": 15,
}


class ExcelParseError(Exception):
    """Raised when Excel parsing fails."""
    pass


class ParsedHolding:
    """Parsed holding record from Excel."""

    def __init__(
        self,
        fund_code: str,
        fund_name: str,
        share_type: str,
        management_company: str,
        fund_account: str,
        platform: str,
        trade_account: str,
        shares: Decimal,
        share_date: date,
        nav: Optional[Decimal],
        nav_date: Optional[date],
        market_value: Optional[Decimal],
        currency: str,
        dividend_mode: Optional[str],
    ):
        self.fund_code = fund_code
        self.fund_name = fund_name
        self.share_type = share_type
        self.management_company = management_company
        self.fund_account = fund_account
        self.platform = platform
        self.trade_account = trade_account
        self.shares = shares
        self.share_date = share_date
        self.nav = nav
        self.nav_date = nav_date
        self.market_value = market_value
        self.currency = currency
        self.dividend_mode = dividend_mode

    @property
    def unique_key(self) -> tuple:
        return (self.fund_code, self.platform, self.fund_account, self.trade_account)


def _parse_date(value: str | None) -> Optional[date]:
    """Parse date string in various formats."""
    if not value:
        return None
    value = str(value).strip()
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y%m%d"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    return None


def _parse_decimal(value: str | None) -> Optional[Decimal]:
    """Parse decimal value from string."""
    if not value:
        return None
    try:
        return Decimal(str(value).strip().replace(",", ""))
    except (InvalidOperation, ValueError):
        return None


def compute_file_hash(file_path: str | Path) -> str:
    """Compute SHA256 hash of a file."""
    h = hashlib.sha256()
    with open(file_path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def validate_headers(ws) -> bool:
    """Validate that row 5 contains expected column headers."""
    row5 = [cell.value for cell in ws[5]]
    expected = ["序号", "基金代码", "基金名称", "份额类别", "基金管理人"]
    for i, exp in enumerate(expected):
        actual = str(row5[i]).strip() if row5[i] else ""
        if actual != exp:
            raise ExcelParseError(
                f"列头验证失败: 第{i+1}列期望'{exp}', 实际'{actual}'"
            )
    return True


def parse_excel(file_path: str | Path) -> tuple[list[ParsedHolding], list[dict], Optional[date]]:
    """Parse fund holdings Excel file.

    Returns:
        (holdings, errors, data_date)
        - holdings: list of successfully parsed holding records
        - errors: list of error dicts with row number and message
        - data_date: the common date from the data, if consistent
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    # Validate headers
    validate_headers(ws)

    holdings: list[ParsedHolding] = []
    errors: list[dict] = []
    data_date: Optional[date] = None

    for row_idx in range(6, ws.max_row + 1):
        row = [ws.cell(row=row_idx, column=c).value for c in range(1, 16)]

        # Skip empty rows or non-data rows
        seq = row[0]
        if seq is None:
            continue
        try:
            float(str(seq))
        except ValueError:
            continue

        fund_code = str(row[1]).strip() if row[1] else None
        if not fund_code:
            errors.append({"row": row_idx, "message": "基金代码为空"})
            continue

        fund_name = str(row[2]).strip() if row[2] else ""
        share_type = str(row[3]).strip() if row[3] else "前收费"
        mgmt_company = str(row[4]).strip() if row[4] else ""
        fund_account = str(row[5]).strip() if row[5] else ""
        platform = str(row[6]).strip() if row[6] else ""
        trade_account = str(row[7]).strip() if row[7] else ""

        shares = _parse_decimal(row[8])
        if shares is None:
            errors.append({"row": row_idx, "message": f"份额解析失败: {row[8]}"})
            continue

        share_date = _parse_date(row[9])
        if share_date is None:
            errors.append({"row": row_idx, "message": f"份额日期解析失败: {row[9]}"})
            continue

        nav = _parse_decimal(row[10])
        nav_dt = _parse_date(row[11])
        market_value = _parse_decimal(row[12])
        currency = str(row[13]).strip() if row[13] else "人民币"
        dividend_mode = str(row[14]).strip() if row[14] else None

        # Track data date (use first valid share_date)
        if data_date is None:
            data_date = share_date

        holdings.append(ParsedHolding(
            fund_code=fund_code,
            fund_name=fund_name,
            share_type=share_type,
            management_company=mgmt_company,
            fund_account=fund_account,
            platform=platform,
            trade_account=trade_account,
            shares=shares,
            share_date=share_date,
            nav=nav,
            nav_date=nav_dt,
            market_value=market_value,
            currency=currency,
            dividend_mode=dividend_mode,
        ))

    wb.close()
    return holdings, errors, data_date
